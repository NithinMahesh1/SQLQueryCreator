import pandas
import os
import pyodbc 
import re
import string
import xlwt 
from xlwt import Workbook 
import xlsxwriter

import openpyxl
from openpyxl import load_workbook

count = 0

def main():
    reading(count)


def reading(count):
    num = 0
    path = os.getcwd() + "\Width Lengths.xlsx"
    data = pandas.read_excel(path, sheet_name='Sheet1')
    sheet2 = pandas.read_excel(path, sheet_name='Sheet2')

    sheet3 = load_workbook('Width Lengths.xlsx')
    sheet3.create_sheet('Sheet3')

    standardLengths = []
    for index, rows in sheet2.iterrows():
        lengths = rows['Standard']
        if pandas.isnull(lengths):
            lengths = "N/A"
            standardLengths.append(lengths)
        else:
            standardLengths.append(lengths)

    items = {}
    for index, row in data.iterrows():
        key = row['ItemType']
        value = row['Width']

        if not (pandas.isnull(value)):
            value = int(value)

        check1 = pandas.isnull(key)
        check2 = pandas.isnull(value)

        if check1 == True and check2 == True:
            continue

        if key == "Finish":
            num = num + 1
            boolean = False
            count = len(items) - 1

            # loop through lengths here then remove up the count from 0 to count
            lengths = []
            increment = 0
            for elem in standardLengths:
                increment = increment + 1
                lengths.append(standardLengths.pop(0))
                # print(str(elem))
                if increment == count:
                    break
                # continue
            
            if len(standardLengths) == 1:
                lengths.append(standardLengths.pop(0))    
            

            query(items,sheet3,num,lengths,count)
            continue

        concat = key, value
        print(concat)

        if check1 == False and check2 == True:
            items[key] = value
            boolean = True
           
        if boolean == True:
            items[key] = value

def query(items,sheet3,num,lengths,count):
    conn = pyodbc.connect('Driver={SQL Server};'
                      'Server=AND692557\SQLEXPRESS;'
                      'Database=12_SP11;'
                      'Trusted_Connection=yes;')

    cursor = conn.cursor()
    itemProperty = []
    for key,value in items.items():
        propertyName = []
        if len(key) > 0 and pandas.isnull(value):
            itemtype = key
        
        # if key and value are not null then save properties
        if len(key) > 0 and not pandas.isnull(value):
            itemProperty.append(key)
            itemProperty.append(value)  

        print(key,value)

    # Get itemtypes source ID first 
    querySource_ID = "SELECT ID FROM innovator.ITEMTYPE WHERE NAME='" +itemtype+ "'"
    print(querySource_ID)

    cursor.execute(querySource_ID)

    # Grab the ID from response
    for row in cursor:
        row = str(row).split("'",1)
        row = row[1]
        row = row.split("',",1)
        ID = row[0]

        print(ID)

    wks = sheet3['Sheet3']
    lastRow = wks.max_row

    # Query for old width lenghts and save those before running change query
    # SELECT COLUMN_WIDTH FROM innovator.PROPERTY WHERE name='property name' AND SOURCE_ID='the id we got'
    compareWidths = []
    for properties in itemProperty:
        lastRow = wks.max_row  
        lastColumn = wks.max_column
        checkRowandColumn = False
        
        if not isinstance(properties, int):
            propertyName = properties

            queryOldLengths = "SELECT COLUMN_WIDTH FROM innovator.PROPERTY WHERE NAME=" + "'" +properties+ "'" " AND SOURCE_ID=" +"'" +ID+ "'"
            cursor.execute(queryOldLengths)
            if cursor.rowcount == 0:
                print("Error in " +propertyName)
                # break
            
            for row in cursor:
                row = str(row).split(",",1)
                row = row[0].split("(",1)
                if str(row[1]) == "None":
                    row = "N/A"
                else:
                    row = int(row[1])     

                compareWidths = row
                
                if lastColumn == 1:
                    wks.cell(row=lastRow, column=1).value = itemtype
                    wks.cell(row=lastRow, column=2).value = properties
                    wks.cell(row=lastRow, column=5).value = row
                    sheet3.save('Width Lengths.xlsx')
                    checkRowandColumn = True
                    
                if checkRowandColumn == False:
                    if num >= 2:
                        emptyRow = ""
                        wks.cell(row=lastRow+1, column=1).value = itemtype

                    wks.cell(row=lastRow+1, column=2).value = properties
                    wks.cell(row=lastRow+1, column=5).value = row
                    sheet3.save('Width Lengths.xlsx')

      
                checkRowandColumn == False

        widthLength = 0
        if isinstance(properties, int):
            widthLength = properties
        
        # UPDATE innovator.PROPERTY SET COLUMN_WIDTH='new int width' WHERE NAME='property name' AND SOURCE_ID='ID variable'
        # Setup second query using source ID and property values 
        if not widthLength == 0 and not propertyName == "":

            # print(compareWidths)
            wks.cell(row=lastRow, column=4).value = lengths[0]

            # intLengths = 0
            stdLengths = lengths[0]

            if lengths[0] == "N/A":
                stdLengths = " IS NULL"
            else:
                intLengths = int(lengths[0])
                stdLengths = "=" +"'" +str(intLengths)+ "'"

            updateQuery = ""
            if lengths[0] == compareWidths:
                updateQuery = "UPDATE innovator.PROPERTY SET COLUMN_WIDTH=" + "'" +str(widthLength)+ "'" + " WHERE NAME=" + "'" +propertyName+ "'" +" AND SOURCE_ID=" + "'" +ID+ "'" +" AND COLUMN_WIDTH" +stdLengths
                wks.cell(row=lastRow, column=6).value = updateQuery
                sheet3.save('Width Lengths.xlsx')
                lastRow = wks.max_row
            else:
                updateQuery = "UPDATE innovator.PROPERTY SET COLUMN_WIDTH=" + "'" +str(widthLength)+ "'" + " WHERE NAME=" + "'" +propertyName+ "'" +" AND SOURCE_ID=" + "'" +ID+ "'" +" AND COLUMN_WIDTH" +str(stdLengths)
                wks.cell(row=lastRow, column=6).value = updateQuery
                sheet3.save('Width Lengths.xlsx')

            checkRowandColumn = True
            lengths.pop(0)
            checkRowandColumn == False
 
    print(num)
               
    items.clear()


main()